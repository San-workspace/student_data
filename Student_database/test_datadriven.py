import requests
import json
import openpyxl


def test_std_data():
    # api
    url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open("C:\\Personal\\API_activity\\stdn_api\\req_json.json", 'r')
    inp_json = file.read()
    req_json = json.loads(inp_json)

    # openpyxl
    book = openpyxl.load_workbook('C:\\Users\\fsn3kor\\PycharmProjects\\pythonProject1\\Student_database\\student_data'
                                  '.xlsx')
    sheet = book['Sheet1']
    rows=sheet.max_row

    for i in range(2, sheet.max_row + 1):
        cell_first_name = sheet.cell(row=i, column=1)
        cell_middle_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_date_of_birth = sheet.cell(row=i, column=4)
        req_json['first_name'] = cell_first_name.value
        req_json['middle_name'] = cell_middle_name.value
        req_json['last_name'] = cell_last_name.value
        req_json['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(url, req_json)
        print(response.text)
        print(response.status_code)
