import openpyxl
import requests


class base:

    def __init__(self, filepath, sheetname):
        global sheet
        global book
        book = openpyxl.load_workbook(filepath)
        sheet = book[sheetname]

    def fetch_col_count(self):
        col = sheet.max_column
        return col

    def fetch_row_count(self):
        row = sheet.max_row
        return row

    def fetch_keynames(self):
        c = sheet.max_column
        l = []
        for i in range(1, c + 1): #iterating column
            cell = sheet.cell(row=1, column=i)
            l.insert(i - 1,cell.value)  # i-1 --> i starting from ,if i=1 its pointing row=1 col=1 and cell.value fetch the row=1 col=1 value to list[i-1] means l[0] index ans so on

        return l

    def update_reqjson_with_data(self, rownumber, jsonRequest, keylist):
        c = sheet.max_column
        for i in range(1, c + 1):
            cell = sheet.cell(row=rownumber, column=i)
            jsonRequest[keylist[i - 1]] = cell.value
        return jsonRequest
